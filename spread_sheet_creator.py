import openpyxl

# Try to load the existing spreadsheet, or create a new one if it doesn't exist
try:
    workbook = openpyxl.load_workbook('existing_spreadsheet.xlsx')
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active

# Define data for the first column
first_column_data = ['League Name', 'Team Name', 'qb', 'rb', 'rb', 'wr', 'wr', 'te', 'kicker', 'defense']

# Append data to the first column
for index, value in enumerate(first_column_data, start=1):
    sheet.cell(row=index, column=1, value=value)

# Define data to append to the other columns
new_column_data = ['New Value 1', 'New Value 2', 'New Value 3']

# Find the next empty column
max_col_index = sheet.max_column
next_col_index = max_col_index + 1

# Append data to other columns
for index, value in enumerate(new_column_data, start=1):
    sheet.cell(row=index, column=next_col_index, value=value)

# Save the changes
workbook.save('existing_spreadsheet.xlsx')
