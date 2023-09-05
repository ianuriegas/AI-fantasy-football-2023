import os
import json
import re
# from qb import get_qb_pytorch_stats
from rb import get_qb_pytorch_stats, get_rb_pytorch_stats, get_wr_pytorch_stats, get_te_pytorch_stats, get_kicker_pytorch_stats, get_defense_pytorch_stats
import openpyxl
# Path to the folder containing JSON files
folder_path = "data/team_data"
player_info_file = "player_info.xlsx"
# Create empty lists for each position


def print_players(position_stats, position_name, max_count):
    count = 0
    for player_name, _ in position_stats:
        print(f"{position_name}: {player_name}")
        count += 1
        if count == max_count:
            break


def get_players(position_stats, position_name, max_count):
    players = []
    count = 0
    for player_name, _ in position_stats:
        players.append((position_name, player_name))
        count += 1
        if count == max_count:
            break
    return players


def print_pretty_json(title, data):
    print(f"{title}:")
    pretty_json = json.dumps(data, indent=4)
    print(pretty_json)


def get_combined_length(*arrays):
    total_length = sum(len(arr) for arr in arrays)
    return total_length


def parse_json_filename(filename):
    pattern = r'TEAM-(\d+)-(\d+)\.json'
    match = re.match(pattern, filename)

    if match:
        league_id = match.group(1)
        team_id = match.group(2)
        return league_id, team_id
    else:
        return None, None


def get_league_and_team_name(league_id, team_id):
    league_data_path = f"data/league_data/LEAGUE-{league_id}-{team_id}.json"
    with open(league_data_path, "r") as json_file:
        data = json.load(json_file)
        league_name = data['settings']['name']
        team_name = data['teams'][team_id-1]['name']
        # print(league_name)
        # print(team_name)
        return league_name, team_name


all_info = []
# Iterate through each JSON file in the folder


def main():
    qb_players = []
    rb_players = []
    wr_players = []
    te_players = []
    defense_players = []
    kicker_players = []

    # Delete the player_info.xlsx file if it already exists
    if os.path.exists(player_info_file):
        os.remove(player_info_file)

    wb = openpyxl.Workbook()

    for filename in os.listdir(folder_path):
        if filename.endswith(".json"):
            file_path = os.path.join(folder_path, filename)
            # print(filename)
            league_id, team_id = parse_json_filename(filename)
            league_name, team_name = get_league_and_team_name(
                league_id, int(team_id))
            with open(file_path, "r") as json_file:
                # print(json_file)

                data = json.load(json_file)
                for player in data:
                    position = player.get("position", "").lower()
                    if position == "qb":
                        qb_players.append(player)
                    elif position == "rb":
                        rb_players.append(player)
                    elif position == "wr":
                        wr_players.append(player)
                    elif position == "te":
                        te_players.append(player)
                    elif position == "defense":
                        defense_players.append(player)
                    elif position == "kicker":
                        kicker_players.append(player)

                # print(json.dumps(rb_players, indent=4))
                print("=================================================================================================================")

                # Getting the sorted data arrays
                qb_stats = get_qb_pytorch_stats(qb_players)
                rb_stats = get_rb_pytorch_stats(rb_players)
                wr_stats = get_wr_pytorch_stats(wr_players)
                te_stats = get_te_pytorch_stats(te_players)
                kicker_stats = get_kicker_pytorch_stats(kicker_players)
                defense_stats = get_defense_pytorch_stats(defense_players)

                print("League Name:", league_name)
                print("Team Name:", team_name)

                qb_players_list = get_players(qb_stats, "QB", 1)
                print(qb_players_list)
                rb_players_list = get_players(rb_stats, "RB", 2)
                print(rb_players_list)
                wr_players_list = get_players(wr_stats, "WR", 2)
                print(wr_players_list)
                te_players_list = get_players(te_stats, "TE", 1)
                print(te_players_list)
                kicker_players_list = get_players(kicker_stats, "Kicker", 1)
                print(kicker_players_list)
                defense_players_list = get_players(defense_stats, "Kicker", 1)
                print(defense_players_list)

                all_players = [
                    qb_players_list,
                    rb_players_list,
                    wr_players_list,
                    te_players_list,
                    kicker_players_list,
                    defense_players_list
                ]
                # Create a new worksheet for each league
                sheet_title = league_name[:31]

                ws = wb.create_sheet(sheet_title)

                # Start from column B (index 2)
                for col_idx, players in enumerate(all_players, start=2):
                    for position, player_name in players:
                        ws.cell(row=1, column=col_idx,
                                value=position)  # Header row
                        ws.cell(row=players.index((position, player_name)) + 2,
                                column=col_idx, value=player_name)  # Data rows

                qb_players = []
                rb_players = []
                wr_players = []
                te_players = []
                kicker_players = []
                defense_players = []
    wb.save(player_info_file)
    print("=================================================================================================================")


if __name__ == "__main__":
    main()
